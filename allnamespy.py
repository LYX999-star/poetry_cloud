# -*- coding: utf-8 -*-
"""
Created on Fri Mar  5 15:35:10 2021

@author: xingyu
"""

from dictionary import get2500,get3500
import openpyxl
     
#dic = dic[:1000]    
poetrycloud = [""]

dic = get3500()

def dfs(n):
    global poetrycloud
    print(n)
    if n <= 0:
        return ;
    print("len::",len(poetrycloud))
    tmp = []
    for _,s in enumerate(poetrycloud):
        
        for _,word in enumerate(dic):
            ts = s
            ts += word
            #print(ts)
            tmp.append(ts)
    poetrycloud = tmp
    n -= 1
    dfs(n)


def poetry2num(poetry):
    res = 0
    for i in range(len(poetry)):
        res *= len(dic)
        res += (dic.index(poetry[i])+1)
    return res

def num2poetry(num):
    res = ""
    while num != 0:
        res += dic[int(num%(len(dic)))-1]
        num //= len(dic)
    return res[::-1]
    

def getAll(n = 1):
    path = 'allnames.xlsx'
    workbook0 = openpyxl.Workbook()
    sheet = workbook0.active
    sheet.title = 'Sheet1'
    #sheet.cell(row=k, column=j+1, value=str(sheet1.cell(i,j).value))

    workbook0.save(path)
    i = 1
    while i < n+1:        
        for j in range(i,i+10000):
            sheet.cell(row=j, column=1, value=num2poetry(j))
        print(i)
        i = i + 10000
        workbook0.save(path)
    
def save(names):
    path = 'allnames3500_3500.xlsx'
    workbook0 = openpyxl.Workbook()
    sheet = workbook0.active
    sheet.title = 'Sheet1'
    #sheet.cell(row=k, column=j+1, value=str(sheet1.cell(i,j).value))
    
    col = 1
    row = 1
    for _,name in enumerate(names):
        
        sheet.cell(row=row, column=col, value=name)
        
        col = (col + 1)
        if col == 16385:
            row = row + 1
            col = 1
        
        
    workbook0.save(path)
    
#dfs(1)
if __name__ == "__main__":
    #getAll(n = 3500*3500)
    dfs(2)
    save(poetrycloud)
"""
        poetry = "床前明月光疑是地上霜举头望明月低头思故乡"#"啊啊啊啊啊啊啊啊啊啊啊啊啊啊啊哎"
        print("1.数字转诗\n2.诗转数字")
        f = int(input("你的选择是："))
        if f == 2:
            poetry = input("输入诗：")
            num = poetry2num(poetry)
            print(int(num))
        if f == 1:
            num = int(input("输入数字："))
            poetry = num2poetry(num)
            print(poetry)
        f = input("Q/q退出")

"""





















